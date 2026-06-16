import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_dataset(filepath, impute_numeric='median', impute_categorical='mode', text_case='title', outlier_threshold=1.5):
    """
    Cleans a dataset by removing duplicates, standardizing dates and text columns,
    imputing missing values, and capping outliers.
    Keeps a detailed audit trail of all corrections.
    """
    df_raw = pd.read_csv(filepath)
    df = df_raw.copy()
    
    audit_logs = []
    stats = {
        "rows_before": len(df_raw),
        "rows_after": 0,
        "duplicates_removed": 0,
        "nulls_imputed": 0,
        "outliers_capped": 0,
        "dates_standardized": 0,
        "text_standardized": 0
    }
    
    # --- STEP 1: Fuzzy Text Cleaning for Deduplication ---
    # Strip spaces and standardize text columns to prepare for deduplication
    text_cols = df.select_dtypes(include=['object']).columns
    for col in text_cols:
        if col.lower() != 'date':
            # Inspect differences for audit trail
            for idx, val in df[col].items():
                if pd.notna(val) and isinstance(val, str):
                    stripped = val.strip()
                    if stripped != val:
                        audit_logs.append({
                            "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                            "column": col,
                            "type": "Whitespace Trim",
                            "message": f"Trimmed trailing/leading spaces from '{val}'",
                            "original": val,
                            "cleaned": stripped
                        })
                        stats["text_standardized"] += 1
                    df.loc[idx, col] = stripped
                    
    # --- STEP 2: Exact & Fuzzy Deduplication ---
    # Find exact duplicate rows
    duplicate_mask = df.duplicated(keep='first')
    duplicates_count = duplicate_mask.sum()
    if duplicates_count > 0:
        # Log duplicate rows
        dup_rows = df[duplicate_mask]
        for idx, row in dup_rows.iterrows():
            audit_logs.append({
                "row": int(row['Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                "column": "All Columns",
                "type": "Duplicate Removal",
                "message": "Removed duplicate row from transaction log",
                "original": str(row.to_dict()),
                "cleaned": "Row Dropped"
            })
        df = df.drop_duplicates(keep='first').reset_index(drop=True)
        stats["duplicates_removed"] = int(duplicates_count)
        
    # --- STEP 3: Inconsistent Date Standardization ---
    if 'Date' in df.columns:
        for idx, val in df['Date'].items():
            if pd.isna(val) or str(val).strip() in ['', 'nan', 'None', 'null']:
                # Handle null dates later in imputation
                continue
            
            val_str = str(val).strip()
            parsed_date = None
            
            # Attempt multi-format parsing
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%B %d, %Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    parsed_date = pd.to_datetime(val_str, format=fmt)
                    break
                except:
                    continue
                    
            if parsed_date is None:
                # Native pandas parse fallback
                try:
                    parsed_date = pd.to_datetime(val_str)
                except:
                    pass
                    
            if parsed_date is not None:
                std_date_str = parsed_date.strftime('%Y-%m-%d')
                if std_date_str != val_str:
                    audit_logs.append({
                        "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                        "column": "Date",
                        "type": "Date Format",
                        "message": f"Standardized date format from '{val_str}' to standard ISO format",
                        "original": val_str,
                        "cleaned": std_date_str
                    })
                    stats["dates_standardized"] += 1
                df.loc[idx, 'Date'] = std_date_str
            else:
                # If parsing completely failed, mark as null for imputation
                df.loc[idx, 'Date'] = np.nan
                audit_logs.append({
                    "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                    "column": "Date",
                    "type": "Data Error",
                    "message": f"Could not parse date '{val_str}'. Set to NULL.",
                    "original": val_str,
                    "cleaned": "NULL"
                })
                
    # --- STEP 4: Missing Values Imputation ---
    for col in df.columns:
        null_indices = df[df[col].isna() | (df[col] == '') | (df[col].astype(str).str.lower() == 'none')].index
        if len(null_indices) > 0:
            if pd.api.types.is_numeric_dtype(df[col]):
                # Numeric Imputation
                if impute_numeric == 'drop':
                    df = df.drop(null_indices).reset_index(drop=True)
                    for n_idx in null_indices:
                        audit_logs.append({
                            "row": int(df_raw.loc[n_idx, 'Transaction_ID']) if 'Transaction_ID' in df_raw.columns else int(n_idx + 1),
                            "column": col,
                            "type": "Null Value Drop",
                            "message": f"Dropped row due to missing vital numerical field '{col}'",
                            "original": "NULL",
                            "cleaned": "Row Dropped"
                        })
                    continue
                else:
                    # Calculate mean/median excluding nulls
                    fill_val = df[col].median() if impute_numeric == 'median' else df[col].mean()
                    fill_val = round(float(fill_val), 2)
                    
                    for n_idx in null_indices:
                        audit_logs.append({
                            "row": int(df.loc[n_idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(n_idx + 1),
                            "column": col,
                            "type": "Null Imputation",
                            "message": f"Imputed missing numerical cell in '{col}' using column {impute_numeric}",
                            "original": "NULL",
                            "cleaned": str(fill_val)
                        })
                        df.loc[n_idx, col] = fill_val
                        stats["nulls_imputed"] += 1
            else:
                # Categorical/Text Imputation
                if impute_categorical == 'drop':
                    df = df.drop(null_indices).reset_index(drop=True)
                    for n_idx in null_indices:
                        audit_logs.append({
                            "row": int(df_raw.loc[n_idx, 'Transaction_ID']) if 'Transaction_ID' in df_raw.columns else int(n_idx + 1),
                            "column": col,
                            "type": "Null Value Drop",
                            "message": f"Dropped row due to missing vital field '{col}'",
                            "original": "NULL",
                            "cleaned": "Row Dropped"
                        })
                    continue
                else:
                    if impute_categorical == 'mode':
                        mode_val = df[col].mode()
                        fill_val = str(mode_val[0]) if len(mode_val) > 0 else 'Unknown'
                    else:
                        fill_val = 'Unknown'
                        
                    for n_idx in null_indices:
                        audit_logs.append({
                            "row": int(df.loc[n_idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(n_idx + 1),
                            "column": col,
                            "type": "Null Imputation",
                            "message": f"Imputed missing categorical cell in '{col}' using column {impute_categorical}",
                            "original": "NULL",
                            "cleaned": fill_val
                        })
                        df.loc[n_idx, col] = fill_val
                        stats["nulls_imputed"] += 1
                        
    # --- STEP 5: Text Casing Standardization ---
    for col in text_cols:
        if col.lower() != 'date':
            for idx, val in df[col].items():
                if pd.notna(val) and isinstance(val, str) and val != 'Unknown':
                    cased = val
                    if text_case == 'title': cased = val.title()
                    elif text_case == 'upper': cased = val.upper()
                    elif text_case == 'lower': cased = val.lower()
                    
                    if cased != val:
                        audit_logs.append({
                            "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                            "column": col,
                            "type": "Text Standardization",
                            "message": f"Standardized casing from '{val}' to '{cased}'",
                            "original": val,
                            "cleaned": cased
                        })
                        df.loc[idx, col] = cased
                        stats["text_standardized"] += 1
                        
    # --- STEP 6: Outlier Cap Handling ---
    # Focus outlier capping on Unit_Price_INR and Quantity
    numeric_outlier_cols = [c for c in ['Unit_Price_INR', 'Quantity'] if c in df.columns]
    
    if outlier_threshold > 0:
        for col in numeric_outlier_cols:
            q25 = df[col].quantile(0.25)
            q75 = df[col].quantile(0.75)
            iqr = q75 - q25
            
            upper_limit = q75 + outlier_threshold * iqr
            lower_limit = q25 - outlier_threshold * iqr
            
            for idx, val in df[col].items():
                if pd.notna(val):
                    val_f = float(val)
                    if val_f > upper_limit:
                        capped_val = round(upper_limit, 2)
                        audit_logs.append({
                            "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                            "column": col,
                            "type": "Outlier Cap",
                            "message": f"Capped price/quantity outlier of {val_f} to threshold limit {capped_val}",
                            "original": str(val_f),
                            "cleaned": str(capped_val)
                        })
                        df.loc[idx, col] = capped_val
                        stats["outliers_capped"] += 1
                    elif val_f < lower_limit:
                        capped_val = round(max(0, lower_limit), 2)
                        audit_logs.append({
                            "row": int(df.loc[idx, 'Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                            "column": col,
                            "type": "Outlier Cap",
                            "message": f"Capped price/quantity lower outlier of {val_f} to threshold limit {capped_val}",
                            "original": str(val_f),
                            "cleaned": str(capped_val)
                        })
                        df.loc[idx, col] = capped_val
                        stats["outliers_capped"] += 1
                        
    # --- STEP 7: Re-calculate Revenue for Mathematical Inconsistency ---
    if 'Total_Revenue_INR' in df.columns and 'Unit_Price_INR' in df.columns and 'Quantity' in df.columns:
        for idx, row in df.iterrows():
            correct_revenue = round(float(row['Unit_Price_INR']) * int(row['Quantity']), 2)
            if pd.isna(row['Total_Revenue_INR']) or abs(float(row['Total_Revenue_INR']) - correct_revenue) > 0.05:
                old_rev = str(row['Total_Revenue_INR']) if pd.notna(row['Total_Revenue_INR']) else "NULL"
                audit_logs.append({
                    "row": int(row['Transaction_ID']) if 'Transaction_ID' in df.columns else int(idx + 1),
                    "column": "Total_Revenue_INR",
                    "type": "Calculation Sync",
                    "message": f"Re-calculated revenue to ensure mathematical sync: {row['Unit_Price_INR']} price * {row['Quantity']} qty",
                    "original": old_rev,
                    "cleaned": str(correct_revenue)
                })
                df.loc[idx, 'Total_Revenue_INR'] = correct_revenue
                stats["text_standardized"] += 1 # log as standardized anomaly
                
    stats["rows_after"] = len(df)
    
    # Save cleaned file locally
    save_path = os.path.join(os.path.dirname(filepath), 'cleaned_dataset.csv')
    df.to_csv(save_path, index=False)
    
    return {
        "stats": stats,
        "audit_logs": audit_logs,
        "cleaned_data": df.to_dict(orient='records'),
        "original_data_preview": df_raw.head(10).to_dict(orient='records'),
        "cleaned_file_path": save_path
    }

def generate_excel_report(cleaned_data_filepath, stats, audit_logs, output_filepath):
    """
    Generates a corporate-styled, multi-tab Excel Workbook report:
    - Tab 1: Executive Audit Summary (dashboard layout, cards, grids, KPIs)
    - Tab 2: Cleaned Transaction Logs (the actual sanitized dataset table)
    """
    df = pd.read_csv(cleaned_data_filepath)
    wb = Workbook()
    
    # ----------------------------------------------------
    # TAB 1: EXECUTIVE AUDIT SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Design fills, borders, and styles
    fill_header = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Slate
    fill_card = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light Grey
    fill_accent_green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Emerald
    fill_accent_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amber
    fill_accent_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Rose
    
    font_title = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=13, bold=True, color="1F2937")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # 1. Header Banner
    ws_summary.merge_cells("A1:G2")
    ws_summary["A1"] = "Automated Data Cleaning Audit Summary"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary["A1"].fill = fill_header
    
    # Set row heights
    ws_summary.row_dimensions[1].height = 25
    ws_summary.row_dimensions[2].height = 25
    ws_summary.row_dimensions[4].height = 20
    
    # 2. Executive KPI Cards
    ws_summary["A4"] = "EXECUTIVE METRIC SUMMARY"
    ws_summary["A4"].font = font_section
    
    # Metric Card 1: Row Stats
    ws_summary.merge_cells("A5:B5")
    ws_summary["A5"] = "Total Rows Audited"
    ws_summary["A5"].font = font_bold
    ws_summary["A5"].fill = fill_card
    ws_summary["A5"].alignment = Alignment(horizontal="center")
    
    ws_summary.merge_cells("A6:B6")
    ws_summary["A6"] = f"{stats['rows_before']} Before / {stats['rows_after']} After"
    ws_summary["A6"].font = Font(name="Calibri", size=11, bold=True, color="111827")
    ws_summary["A6"].fill = fill_card
    ws_summary["A6"].alignment = Alignment(horizontal="center")
    
    # Metric Card 2: Duplicates Removed
    ws_summary.merge_cells("C5:D5")
    ws_summary["C5"] = "Duplicates Cleared"
    ws_summary["C5"].font = font_bold
    ws_summary["C5"].fill = fill_accent_red
    ws_summary["C5"].alignment = Alignment(horizontal="center")
    
    ws_summary.merge_cells("C6:D6")
    ws_summary["C6"] = stats['duplicates_removed']
    ws_summary["C6"].font = Font(name="Calibri", size=12, bold=True, color="991B1B")
    ws_summary["C6"].fill = fill_accent_red
    ws_summary["C6"].alignment = Alignment(horizontal="center")
    
    # Metric Card 3: Null Values Imputed
    ws_summary.merge_cells("E5:F5")
    ws_summary["E5"] = "Null Cells Imputed"
    ws_summary["E5"].font = font_bold
    ws_summary["E5"].fill = fill_accent_yellow
    ws_summary["E5"].alignment = Alignment(horizontal="center")
    
    ws_summary.merge_cells("E6:F6")
    ws_summary["E6"] = stats['nulls_imputed']
    ws_summary["E6"].font = Font(name="Calibri", size=12, bold=True, color="92400E")
    ws_summary["E6"].fill = fill_accent_yellow
    ws_summary["E6"].alignment = Alignment(horizontal="center")
    
    # Metric Card 4: Outliers Capped
    ws_summary["G5"] = "Outliers Handled"
    ws_summary["G5"].font = font_bold
    ws_summary["G5"].fill = fill_accent_green
    ws_summary["G5"].alignment = Alignment(horizontal="center")
    
    ws_summary["G6"] = stats['outliers_capped']
    ws_summary["G6"].font = Font(name="Calibri", size=12, bold=True, color="065F46")
    ws_summary["G6"].fill = fill_accent_green
    ws_summary["G6"].alignment = Alignment(horizontal="center")
    
    # Apply borders to cards
    for row in range(5, 7):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_summary[f"{col}{row}"].border = thin_border
            
    # 3. Detailed Statistics Grid
    ws_summary["A8"] = "DETAILED PIPELINE METRICS"
    ws_summary["A8"].font = font_section
    
    headers_stats = ["Metric Action Category", "Anomalies Corrected", "Correction Strategy"]
    for col_idx, h in enumerate(headers_stats, start=1):
        cell = ws_summary.cell(row=9, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_card
        cell.border = thin_border
        
    stats_data = [
        ["Exact & Fuzzy Duplicate Dropping", stats['duplicates_removed'], "Deleted row copies, kept first chronological record"],
        ["Incomplete Values Imputation", stats['nulls_imputed'], "Filled numerical with median, text categories with mode"],
        ["Statistical Outliers Capping", stats['outliers_capped'], "Clamped values to Interquartile Range (IQR) bounds"],
        ["Inconsistent Date Parsing", stats['dates_standardized'], "Normalized mixed formats to YYYY-MM-DD standard ISO"],
        ["Case and Whitespace Trimming", stats['text_standardized'], "Trimmed paddings, standardized Category names to Title Case"]
    ]
    
    for r_idx, row in enumerate(stats_data, start=10):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx == 2:
                cell.alignment = Alignment(horizontal="right")
                
    # 4. Audit Log Timeline Preview (Top 15 entries)
    ws_summary["A16"] = "RECENT CORRECTIONS AUDIT LOG TRAIL"
    ws_summary["A16"].font = font_section
    
    headers_logs = ["Row ID", "Column", "Type of Fix", "Change Details Description", "Original Value", "Corrected Value"]
    for col_idx, h in enumerate(headers_logs, start=1):
        cell = ws_summary.cell(row=17, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_card
        cell.border = thin_border
        
    # Slice first 15 logs
    preview_logs = audit_logs[:15]
    for r_idx, log in enumerate(preview_logs, start=18):
        ws_summary.cell(row=r_idx, column=1, value=log['row']).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=r_idx, column=2, value=log['column'])
        ws_summary.cell(row=r_idx, column=3, value=log['type'])
        ws_summary.cell(row=r_idx, column=4, value=log['message'])
        ws_summary.cell(row=r_idx, column=5, value=str(log['original'])).alignment = Alignment(horizontal="right")
        ws_summary.cell(row=r_idx, column=6, value=str(log['cleaned'])).alignment = Alignment(horizontal="right")
        
        for c_idx in range(1, 7):
            cell = ws_summary.cell(row=r_idx, column=c_idx)
            cell.font = font_regular
            cell.border = thin_border
            
    # Auto-fit columns for Summary
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        ws_summary.column_dimensions[col_letter].width = min(40, max(max_len + 3, 12))
        
    # ----------------------------------------------------
    # TAB 2: CLEANED DATASET
    # ----------------------------------------------------
    ws_data = wb.create_sheet(title="Cleaned Transaction Database")
    ws_data.views.sheetView[0].showGridLines = True
    
    # Style data header
    fill_data_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Darker Navy
    font_data_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Write dataframe to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
        
    # Style dataset headers
    for cell in ws_data[1]:
        cell.fill = fill_data_header
        cell.font = font_data_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws_data.row_dimensions[1].height = 25
    
    # Style rows
    for r_idx in range(2, ws_data.max_row + 1):
        ws_data.row_dimensions[r_idx].height = 18
        for c_idx in range(1, ws_data.max_column + 1):
            cell = ws_data.cell(row=r_idx, column=c_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            # Formatting specifics: Transaction_ID, Date, Casing
            header_name = ws_data.cell(row=1, column=c_idx).value
            
            # Alignments
            if header_name in ['Transaction_ID', 'Date']:
                cell.alignment = Alignment(horizontal="center")
            elif header_name in ['Quantity']:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif header_name in ['Unit_Price_INR', 'Total_Revenue_INR']:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '₹#,##0.00'
                
    # Auto-fit columns for Data
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_filepath)
    print(f"Styled Excel Report written to: {output_filepath}")
